#!/usr/bin/env python3
"""Prepare real GEO processed datasets into standardized matrices for analysis."""

from __future__ import annotations

import argparse
import csv
import gzip
import math
import re
import shutil
import tempfile
from collections import defaultdict
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd


def clean_gene_name(symbol: object, fallback: object) -> str | None:
    text = ""
    for candidate in [symbol, fallback]:
        if candidate is None or (isinstance(candidate, float) and math.isnan(candidate)):
            continue
        text = str(candidate).strip()
        if text:
            break
    if not text:
        return None
    if text in {"NA", "nan", "NaN", "---", "N/A"}:
        if fallback is None or (isinstance(fallback, float) and math.isnan(fallback)):
            return None
        text = str(fallback).strip()
    text = text.split(".")[0] if text.startswith("ENSG") else text
    return text or None


def split_genes(text: object) -> list[str]:
    if text is None or (isinstance(text, float) and math.isnan(text)):
        return []
    raw = str(text).strip()
    if not raw or raw in {"NA", "nan", "NaN", "---"}:
        return []
    parts = re.split(r"[;,/]", raw)
    return [p.strip() for p in parts if p.strip() and p.strip() not in {"NA", "nan", "NaN", "---"}]


def aggregate_matrix_by_gene(df: pd.DataFrame, gene_col: str, sample_cols: list[str]) -> pd.DataFrame:
    tmp = df[[gene_col] + sample_cols].copy()
    tmp = tmp.dropna(subset=[gene_col])
    tmp = tmp[tmp[gene_col].astype(str).str.len() > 0]
    return tmp.groupby(gene_col, sort=True)[sample_cols].mean().sort_index()


def parse_geo_series_matrix(path: Path) -> dict[str, list[str]]:
    meta: dict[str, list[str]] = {}
    with gzip.open(path, "rt") as handle:
        for line in handle:
            if line.startswith("!series_matrix_table_begin"):
                break
            if not line.startswith("!Sample_"):
                continue
            fields = next(csv.reader([line.rstrip("\n")], delimiter="\t"))
            key = fields[0]
            values = [value.strip().strip('"') for value in fields[1:]]
            meta.setdefault(key, []).append(values)
    return meta


def build_characteristics(meta: dict[str, list[str]]) -> list[dict[str, str]]:
    titles = meta.get("!Sample_title", [[]])[0]
    sample_count = len(titles)
    rows = [dict(title=titles[idx]) for idx in range(sample_count)]
    for char_values in meta.get("!Sample_characteristics_ch1", []):
        for idx, value in enumerate(char_values):
            if idx >= sample_count:
                continue
            if ":" in value:
                key, val = value.split(":", 1)
                rows[idx][key.strip().lower().replace("/", "_").replace(" ", "_")] = val.strip()
    return rows


def prepare_gse123976_expression(base_dir: Path, out_dir: Path) -> None:
    xlsx = base_dir / "GSE123976_HF.v.NF_DESeq2.xlsx"
    df = pd.read_excel(xlsx, sheet_name="Unfiltered")
    sample_cols = [c for c in df.columns if re.fullmatch(r"\d+X\d+", str(c))]
    df["gene_id"] = [clean_gene_name(sym, ens) for sym, ens in zip(df["external_gene_name"], df["ensembl_gene_id"])]
    expr = aggregate_matrix_by_gene(df, "gene_id", sample_cols)
    expr.to_csv(out_dir / "gse123976_expression.tsv", sep="\t")

    de_cols = [
        "gene_id",
        "ensembl_gene_id",
        "baseMean",
        "log2FoldChange",
        "lfcSE",
        "stat",
        "pvalue",
        "padj",
        "chromosome_name",
        "start_position",
        "end_position",
    ]
    de = df[de_cols].copy().dropna(subset=["gene_id"])
    de = de.sort_values(["padj", "pvalue", "baseMean"], ascending=[True, True, False], na_position="last")
    de = de.drop_duplicates(subset=["gene_id"], keep="first")
    de.to_csv(out_dir / "gse123976_expression_de.tsv", sep="\t", index=False)

    meta = parse_geo_series_matrix(base_dir / "GSE123976-GPL11154_series_matrix.txt.gz")
    rows = []
    for row in build_characteristics(meta):
        sample_id = row["title"].split(":", 1)[0].strip()
        diagnosis = row.get("diagnosis", "")
        rows.append(
            {
                "sample_id": sample_id,
                "phenotype": "HF" if diagnosis == "HF" else "NF",
                "diagnosis": diagnosis,
                "age": row.get("age", ""),
                "sex": row.get("sex", ""),
                "race": row.get("race", ""),
            }
        )
    pd.DataFrame(rows).to_csv(out_dir / "gse123976_expression_metadata.tsv", sep="\t", index=False)


def prepare_gse123976_methylation(base_dir: Path, out_dir: Path) -> None:
    gz_xlsx = base_dir / "GSE123976_Table_S1_DMCsv2.xlsx.gz"
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_xlsx = Path(tmpdir) / "gse123976_dmc.xlsx"
        with gzip.open(gz_xlsx, "rb") as src, tmp_xlsx.open("wb") as dst:
            shutil.copyfileobj(src, dst)
        df = pd.read_excel(tmp_xlsx, sheet_name="DMCs - Q < 0.05")

    sample_cols = [c for c in df.columns if str(c).startswith("perc.mC_")]
    promoter = df["Annotation"].astype(str).str.contains("promoter", case=False, na=False).copy()
    df = df.loc[promoter].copy()
    rename = {col: str(col).replace("perc.mC_", "") for col in sample_cols}
    df = df.rename(columns=rename)
    sample_cols = [rename[c] for c in sample_cols]

    rows = []
    for _, row in df.iterrows():
        genes = set(split_genes(row.get("annot.symbol"))) | set(split_genes(row.get("annot.symbol.1")))
        if not genes:
            continue
        values = [float(row[col]) / 100.0 if pd.notna(row[col]) else np.nan for col in sample_cols]
        for gene in genes:
            out = {"gene_id": gene}
            for col, val in zip(sample_cols, values):
                out[col] = val
            out["meth_diff"] = float(row["meth.diff"]) / 100.0 if pd.notna(row["meth.diff"]) else np.nan
            out["qvalue"] = row.get("qvalue", np.nan)
            rows.append(out)

    expanded = pd.DataFrame(rows)
    matrix = expanded.groupby("gene_id", sort=True)[sample_cols].mean().sort_index()
    matrix.to_csv(out_dir / "gse123976_methylation.tsv", sep="\t")

    summary = expanded.groupby("gene_id", sort=True).agg(
        mean_meth_diff=("meth_diff", "mean"),
        min_qvalue=("qvalue", "min"),
        promoter_dmc_count=("gene_id", "size"),
    ).reset_index()
    summary.to_csv(out_dir / "gse123976_methylation_discovery.tsv", sep="\t", index=False)


def prepare_gse116250(base_dir: Path, out_dir: Path) -> None:
    expr = pd.read_csv(base_dir / "GSE116250_rpkm.txt.gz", sep="\t", compression="gzip")
    sample_cols = [c for c in expr.columns if c not in {"Gene", "Common_name"}]
    expr["gene_id"] = [clean_gene_name(sym, ens) for sym, ens in zip(expr["Common_name"], expr["Gene"])]
    matrix = aggregate_matrix_by_gene(expr, "gene_id", sample_cols)
    matrix.to_csv(out_dir / "gse116250_expression.tsv", sep="\t")

    meta = parse_geo_series_matrix(base_dir / "GSE116250_series_matrix.txt.gz")
    rows = []
    for row in build_characteristics(meta):
        disease = row.get("disease", "")
        phenotype = {"non-failing": "NF", "dilated cardiomyopathy": "DCM", "ischemic cardiomyopathy": "ICM"}.get(disease, disease)
        rows.append(
            {
                "sample_id": row["title"],
                "phenotype": phenotype,
                "disease": disease,
                "age": row.get("age", ""),
                "sex": row.get("sex", ""),
                "tissue": row.get("tissue", ""),
            }
        )
    pd.DataFrame(rows).to_csv(out_dir / "gse116250_metadata.tsv", sep="\t", index=False)


def build_gse197670_metadata(base_dir: Path) -> pd.DataFrame:
    rows = []
    for name in ["GSE197670-GPL13534_series_matrix.txt.gz", "GSE197670-GPL21145_series_matrix.txt.gz"]:
        meta = parse_geo_series_matrix(base_dir / name)
        for row in build_characteristics(meta):
            sample_id = row["title"].split(":", 1)[0].strip()
            lvad = row.get("lvad_status", "")
            etiology = row.get("hf_etiology", "")
            phenotype = "NF" if lvad == "Non-Failing" else ("pre_LVAD" if lvad == "pre-LVAD" else "post_LVAD")
            rows.append(
                {
                    "sample_id": sample_id,
                    "phenotype": phenotype,
                    "lvad_status": lvad,
                    "hf_etiology": etiology,
                    "age": row.get("age", ""),
                    "sex": row.get("gender", ""),
                    "race": row.get("race_ethnicity", ""),
                }
            )
    out = pd.DataFrame(rows)
    out["sample_number"] = out["sample_id"].str.extract(r"(\d+)").astype(int)
    out = out.sort_values("sample_number").drop(columns=["sample_number"])
    return out


def build_gse197670_probe_map(family_soft: Path) -> dict[str, set[str]]:
    mapping: dict[str, set[str]] = {}
    with gzip.open(family_soft, "rt", errors="ignore") as handle:
        in_table = False
        header: list[str] | None = None
        col_idx: dict[str, int] = {}
        for raw in handle:
            line = raw.rstrip("\n")
            if line == "!platform_table_begin":
                in_table = True
                header = None
                continue
            if not in_table:
                continue
            if line == "!platform_table_end":
                break
            parts = line.split("\t")
            if header is None:
                header = parts
                col_idx = {name: idx for idx, name in enumerate(header)}
                continue
            probe_id = parts[col_idx["ID"]]
            genes = split_genes(parts[col_idx.get("UCSC_RefGene_Name", -1)])
            if not genes:
                continue
            groups = set(split_genes(parts[col_idx.get("UCSC_RefGene_Group", -1)]))
            regulator = parts[col_idx.get("Regulator", -1)] if "Regulator" in col_idx else ""
            promoter_related = bool(groups.intersection({"TSS1500", "TSS200", "5'UTR", "1stExon"})) or ("Promoter_Associated" in regulator)
            if promoter_related:
                mapping[probe_id] = set(genes)
    return mapping


def prepare_gse197670(base_dir: Path, out_dir: Path) -> None:
    metadata = build_gse197670_metadata(base_dir)
    metadata.to_csv(out_dir / "gse197670_metadata.tsv", sep="\t", index=False)

    probe_map = build_gse197670_probe_map(base_dir / "GSE197670_family.soft.gz")
    workbook = openpyxl.load_workbook(base_dir / "GSE197670_DNAmeth_beta_matrix.xlsx", read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    header = [str(x) if x is not None else "" for x in next(rows)]
    sample_cols = header[1:]

    sums: dict[str, np.ndarray] = defaultdict(lambda: np.zeros(len(sample_cols), dtype=float))
    counts: dict[str, np.ndarray] = defaultdict(lambda: np.zeros(len(sample_cols), dtype=float))

    for row in rows:
        probe_id = row[0]
        if probe_id not in probe_map:
            continue
        values = np.array([float(v) if v is not None else np.nan for v in row[1:]], dtype=float)
        valid = ~np.isnan(values)
        if not valid.any():
            continue
        for gene in probe_map[probe_id]:
            sums[gene][valid] += values[valid]
            counts[gene][valid] += 1

    gene_rows = {}
    for gene, value_sums in sums.items():
        value_counts = counts[gene]
        gene_rows[gene] = np.divide(value_sums, value_counts, out=np.full_like(value_sums, np.nan), where=value_counts > 0)

    matrix = pd.DataFrame.from_dict(gene_rows, orient="index", columns=sample_cols).sort_index()
    matrix.to_csv(out_dir / "gse197670_methylation.tsv", sep="\t")

    if not metadata.empty:
        by_pheno = dict(zip(metadata["sample_id"], metadata["phenotype"]))
        for phenotype, filename in [("pre_LVAD", "gse197670_methylation_pre_lvad.tsv"), ("NF", "gse197670_methylation_nonfailing.tsv")]:
            keep = [col for col in matrix.columns if by_pheno.get(col) == phenotype]
            if keep:
                matrix[keep].to_csv(out_dir / filename, sep="\t")


def main() -> None:
    ap = argparse.ArgumentParser(description="Prepare real GEO processed inputs for downstream analysis")
    ap.add_argument("--data-root", type=Path, default=Path("data/ncbi"))
    ap.add_argument("--out-dir", type=Path, default=Path("results/real_processed/inputs"))
    args = ap.parse_args()

    args.out_dir.mkdir(parents=True, exist_ok=True)
    prepare_gse123976_expression(args.data_root / "GSE123976", args.out_dir)
    prepare_gse123976_methylation(args.data_root / "GSE123976", args.out_dir)
    prepare_gse116250(args.data_root / "GSE116250", args.out_dir)
    prepare_gse197670(args.data_root / "GSE197670", args.out_dir)


if __name__ == "__main__":
    main()